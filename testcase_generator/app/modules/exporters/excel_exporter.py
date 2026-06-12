"""
Excel格式导出器 - 增强版
支持功能测试用例和API测试用例导出
"""
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """导出测试用例为Excel格式"""

    @staticmethod
    def export(test_cases: List[Dict], output_path: str):
        """导出测试用例到Excel文件"""
        wb = Workbook()
        
        # 判断用例类型
        if test_cases and 'method' in test_cases[0]:
            # API测试用例
            ExcelExporter._export_api_cases(wb, test_cases)
        else:
            # 功能测试用例
            ExcelExporter._export_functional_cases(wb, test_cases)
        
        wb.save(output_path)

    @staticmethod
    def _export_functional_cases(wb: Workbook, test_cases: List[Dict]):
        """导出功能测试用例"""
        ws = wb.active
        ws.title = "功能测试用例"

        # 设置表头
        headers = ['用例ID', '用例类型', '用例名称', '描述', '优先级', '前置条件', '测试步骤', '预期结果', '测试数据', '关联需求']
        ExcelExporter._write_header(ws, headers)

        # 写入数据
        for row_idx, tc in enumerate(test_cases, 2):
            # 处理测试步骤
            steps_text = ''
            if isinstance(tc.get('test_steps'), list):
                steps_text = '\n'.join([f"{s.get('step', i+1)}. {s.get('action', '')} → {s.get('expected', '')}" 
                                       for i, s in enumerate(tc.get('test_steps', []))])
            elif tc.get('test_steps'):
                steps_text = str(tc.get('test_steps', ''))

            # 处理测试数据
            test_data_text = ''
            if isinstance(tc.get('test_data'), dict):
                test_data_text = '\n'.join([f"{k}: {v}" for k, v in tc.get('test_data', {}).items()])
            elif tc.get('test_data'):
                test_data_text = str(tc.get('test_data', ''))

            ws.cell(row=row_idx, column=1, value=tc.get('id', ''))
            ws.cell(row=row_idx, column=2, value=tc.get('type', ''))
            ws.cell(row=row_idx, column=3, value=tc.get('name', '')[:50])
            ws.cell(row=row_idx, column=4, value=tc.get('description', '')[:100])
            ws.cell(row=row_idx, column=5, value=tc.get('priority', 'P2'))
            ws.cell(row=row_idx, column=6, value=tc.get('preconditions', ''))
            ws.cell(row=row_idx, column=7, value=steps_text)
            ws.cell(row=row_idx, column=8, value=tc.get('expected', ''))
            ws.cell(row=row_idx, column=9, value=test_data_text)
            ws.cell(row=row_idx, column=10, value=tc.get('requirement', '')[:50])

            # 设置样式
            for col in range(1, 11):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                # 根据优先级设置颜色
                if col == 5:
                    priority = tc.get('priority', 'P2')
                    if priority == 'P0':
                        cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                    elif priority == 'P1':
                        cell.fill = PatternFill(start_color='FFE066', end_color='FFE066', fill_type='solid')

        # 设置列宽
        column_widths = [12, 15, 30, 40, 8, 25, 40, 30, 25, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    @staticmethod
    def _export_api_cases(wb: Workbook, test_cases: List[Dict]):
        """导出API测试用例"""
        ws = wb.active
        ws.title = "API测试用例"

        # 设置表头
        headers = ['用例ID', '用例名称', '请求方法', '请求URL', '请求头', '请求参数', '请求体', '预期状态码', '预期响应', '优先级', '标签', '描述']
        ExcelExporter._write_header(ws, headers)

        # 写入数据
        for row_idx, tc in enumerate(test_cases, 2):
            ws.cell(row=row_idx, column=1, value=tc.get('id', ''))
            ws.cell(row=row_idx, column=2, value=tc.get('name', '')[:50])
            ws.cell(row=row_idx, column=3, value=tc.get('method', ''))
            ws.cell(row=row_idx, column=4, value=tc.get('url', ''))
            
            # 处理JSON格式数据
            headers_text = str(tc.get('headers', {}))
            params_text = str(tc.get('params', {}))
            body_text = str(tc.get('body', {}))
            expected_response_text = str(tc.get('expected_response', {}))
            tags_text = ','.join(tc.get('tags', [])) if isinstance(tc.get('tags'), list) else str(tc.get('tags', ''))

            ws.cell(row=row_idx, column=5, value=headers_text)
            ws.cell(row=row_idx, column=6, value=params_text)
            ws.cell(row=row_idx, column=7, value=body_text)
            ws.cell(row=row_idx, column=8, value=tc.get('expected_status', ''))
            ws.cell(row=row_idx, column=9, value=expected_response_text)
            ws.cell(row=row_idx, column=10, value=tc.get('priority', 'P2'))
            ws.cell(row=row_idx, column=11, value=tags_text)
            ws.cell(row=row_idx, column=12, value=tc.get('description', '')[:50])

            # 设置样式
            for col in range(1, 13):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if col == 10:
                    priority = tc.get('priority', 'P2')
                    if priority == 'P0':
                        cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                    elif priority == 'P1':
                        cell.fill = PatternFill(start_color='FFE066', end_color='FFE066', fill_type='solid')

        # 设置列宽
        column_widths = [12, 35, 8, 25, 20, 20, 20, 12, 25, 8, 20, 35]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    @staticmethod
    def _write_header(ws, headers: List[str]):
        """写入表头"""
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin_border = Border(
            left=Side(style='thin', color='B4B4B4'),
            right=Side(style='thin', color='B4B4B4'),
            top=Side(style='thin', color='B4B4B4'),
            bottom=Side(style='thin', color='B4B4B4')
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border