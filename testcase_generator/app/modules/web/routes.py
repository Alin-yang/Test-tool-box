"""
Web路由
"""
import os
import uuid
import json
import base64
import hashlib
import urllib.parse
from datetime import datetime, timedelta
from typing import Dict, List
from flask import Blueprint, request, jsonify, send_file, render_template
from werkzeug.utils import secure_filename
from app.modules.document_parser import WordParser, PDFParser, MarkdownParser
from app.modules.testcase_generator import FunctionalGenerator, APIGenerator
from app.modules.exporters import XMindExporter, ExcelExporter, ApifoxExporter
from app.modules.data_generator import DataGenerator
from app.modules.mock_generator import MockGenerator

web_bp = Blueprint('web', __name__)

# 使用绝对路径
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'outputs')
ALLOWED_EXTENSIONS = {'docx', 'doc', 'pdf', 'md', 'txt'}

# 全局变量存储服务器配置
server_config = None

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@web_bp.route('/')
def index():
    return render_template('index.html')


@web_bp.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '没有文件上传'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        filepath = os.path.join(UPLOAD_FOLDER, unique_filename)
        file.save(filepath)

        file_ext = filename.rsplit('.', 1)[1].lower()
        parsed_data = parse_document(filepath, file_ext)

        return jsonify({
            'success': True,
            'filename': filename,
            'filepath': unique_filename,
            'data': parsed_data
        })

    return jsonify({'error': '不支持的文件格式'}), 400


@web_bp.route('/api/generate', methods=['POST'])
def generate_testcases():
    data = request.get_json()
    doc_type = data.get('docType', 'requirement')
    parsed_data = data.get('parsedData', {})
    base_url = data.get('baseUrl', '')

    try:
        if doc_type == 'requirement':
            requirements = parsed_data.get('requirements', [])
            generator = FunctionalGenerator()
            test_cases = generator.generate(requirements)
        else:
            api_info = parsed_data.get('apiInfo', [])
            generator = APIGenerator()
            test_cases = generator.generate(api_info, base_url)

        # 保存服务器配置供后续导出使用
        global server_config
        if doc_type == 'api' and base_url:
            server_config = [{'url': base_url, 'description': '测试环境'}]
        else:
            server_config = None

        return jsonify({
            'success': True,
            'testCases': test_cases,
            'count': len(test_cases)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@web_bp.route('/api/generate-with-progress', methods=['POST'])
def generate_testcases_with_progress():
    """带进度的用例生成"""
    data = request.get_json()
    doc_type = data.get('docType', 'requirement')
    parsed_data = data.get('parsedData', {})
    base_url = data.get('baseUrl', '')
    progress = {'step': '准备中...', 'percent': 0}

    try:
        if doc_type == 'requirement':
            requirements = parsed_data.get('requirements', [])
            progress['step'] = '开始生成测试用例...'
            progress['percent'] = 10
            
            generator = FunctionalGenerator()
            progress['step'] = '正在分析等价类...'
            progress['percent'] = 30
            test_cases = generator.generate(requirements)
            
            progress['step'] = '用例生成完成'
            progress['percent'] = 90
        else:
            api_info = parsed_data.get('apiInfo', [])
            progress['step'] = '开始生成接口用例...'
            progress['percent'] = 20
            
            generator = APIGenerator()
            test_cases = generator.generate(api_info, base_url)
            
            progress['step'] = '用例生成完成'
            progress['percent'] = 90

        # 保存服务器配置
        global server_config
        if doc_type == 'api' and base_url:
            server_config = [{'url': base_url, 'description': '测试环境'}]
        else:
            server_config = None

        progress['percent'] = 100
        return jsonify({
            'success': True,
            'testCases': test_cases,
            'count': len(test_cases),
            'progress': progress
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e), 'progress': progress}), 500


@web_bp.route('/api/export', methods=['POST'])
def export_testcases():
    global server_config
    data = request.get_json()
    test_cases = data.get('testCases', [])
    export_format = data.get('format', 'excel')
    doc_type = data.get('docType', 'functional')

    if not test_cases:
        return jsonify({'error': '没有测试用例可导出'}), 400

    extension_map = {
        'xmind': 'xmind',
        'excel': 'xlsx',
        'apifox': 'json',
        'openapi': 'json',
        'apifox-json': 'json',
        'postman': 'json'
    }
    file_ext = extension_map.get(export_format, 'xlsx')
    
    output_filename = f"{uuid.uuid4().hex}_{doc_type}.{file_ext}"
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        if export_format == 'xmind':
            XMindExporter.export(test_cases, output_path)
        elif export_format == 'excel':
            ExcelExporter.export(test_cases, output_path)
        elif export_format == 'apifox' and doc_type == 'api':
            ApifoxExporter.export(test_cases, output_path, server_config)
        elif export_format == 'openapi' and doc_type == 'api':
            ApifoxExporter.export_openapi(test_cases, output_path, server_config)
        elif export_format == 'apifox-json' and doc_type == 'api':
            ApifoxExporter.export_apifox_json(test_cases, output_path, server_config)
        elif export_format == 'postman' and doc_type == 'api':
            ApifoxExporter.export_postman_collection(test_cases, output_path, server_config)
        else:
            ExcelExporter.export(test_cases, output_path)

        return jsonify({
            'success': True,
            'downloadUrl': f'/api/download/{output_filename}',
            'filename': output_filename
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@web_bp.route('/api/download/<filename>')
def download_file(filename):
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': '文件不存在'}), 404


@web_bp.route('/api/generate-data', methods=['POST'])
def generate_test_data():
    data = request.get_json()
    data_type = data.get('type', 'name')
    count = data.get('count', 1)

    try:
        result = []
        for _ in range(count):
            if data_type == 'name':
                result.append(DataGenerator.generate_name())
            elif data_type == 'phone':
                result.append(DataGenerator.generate_phone())
            elif data_type == 'email':
                result.append(DataGenerator.generate_email())
            elif data_type == 'id_card':
                result.append(DataGenerator.generate_id_card())
            elif data_type == 'ip':
                result.append(DataGenerator.generate_ip())
            elif data_type == 'address':
                result.append(DataGenerator.generate_address())
            elif data_type == 'mac':
                result.append(DataGenerator.generate_mac_address())
            elif data_type == 'url':
                result.append(DataGenerator.generate_url())
            elif data_type == 'string':
                result.append(DataGenerator.generate_string())
            elif data_type == 'int':
                result.append(DataGenerator.generate_int())
            elif data_type == 'float':
                result.append(DataGenerator.generate_float())
            elif data_type == 'date':
                result.append(DataGenerator.generate_date())
            elif data_type == 'datetime':
                result.append(DataGenerator.generate_datetime())
            elif data_type == 'user':
                result.append(DataGenerator.generate_user())
            elif data_type == 'boundary_string':
                result.append(DataGenerator.generate_boundary_string())
            elif data_type == 'boundary_number':
                result.append(DataGenerator.generate_boundary_number())
            else:
                result.append(DataGenerator.generate_string())

        return jsonify({'success': True, 'data': result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@web_bp.route('/api/generate-mock', methods=['POST'])
def generate_mock():
    data = request.get_json()
    parsed_data = data.get('parsedData', {})
    api_info = parsed_data.get('apiInfo', [])

    try:
        mock_data = MockGenerator.generate_mock_data(api_info)
        return jsonify({'success': True, 'data': mock_data})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@web_bp.route('/api/mock-templates', methods=['GET'])
def get_mock_templates():
    """获取Mock模板列表"""
    try:
        templates = MockGenerator.get_templates()
        return jsonify({'success': True, 'data': templates})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@web_bp.route('/api/generate-mock-by-template', methods=['POST'])
def generate_mock_by_template():
    """根据模板生成Mock数据"""
    data = request.get_json()
    template_key = data.get('template', 'user')
    
    try:
        result = MockGenerator.generate_by_template(template_key)
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@web_bp.route('/api/export-mock', methods=['POST'])
def export_mock():
    data = request.get_json()
    mock_data = data.get('mockData', [])

    if not mock_data:
        return jsonify({'error': '没有Mock数据可导出'}), 400

    output_filename = f"{uuid.uuid4().hex}_mock.json"
    output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(mock_data, f, ensure_ascii=False, indent=2)

        return jsonify({
            'success': True,
            'downloadUrl': f'/api/download/{output_filename}',
            'filename': output_filename
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@web_bp.route('/api/review-cases', methods=['POST'])
def review_cases():
    data = request.get_json()
    test_cases = data.get('testCases', [])

    if not test_cases:
        return jsonify({'error': '没有测试用例可分析'}), 400

    try:
        # 统计分析
        total = len(test_cases)
        p0_count = sum(1 for tc in test_cases if tc.get('priority') == 'P0')
        p1_count = sum(1 for tc in test_cases if tc.get('priority') == 'P1')
        p2_count = sum(1 for tc in test_cases if tc.get('priority') == 'P2')
        
        # 按场景分类统计
        scenarios = {}
        for tc in test_cases:
            scenario = tc.get('scenario', 'unknown')
            scenarios[scenario] = scenarios.get(scenario, 0) + 1
        
        # 检测重复用例
        seen_names = {}
        duplicates = []
        for tc in test_cases:
            name = tc.get('name', '')
            if name in seen_names:
                duplicates.append({'name': name, 'count': seen_names[name] + 1})
            seen_names[name] = seen_names.get(name, 0) + 1

        result = {
            'total': total,
            'p0_count': p0_count,
            'p1_count': p1_count,
            'p2_count': p2_count,
            'p0_percent': round(p0_count / total * 100, 2) if total > 0 else 0,
            'p1_percent': round(p1_count / total * 100, 2) if total > 0 else 0,
            'p2_percent': round(p2_count / total * 100, 2) if total > 0 else 0,
            'scenarios': scenarios,
            'duplicate_count': len([d for d in duplicates if d['count'] > 1]),
            'duplicates': duplicates[:10],
            'coverage_analysis': {
                'has_normal': any(tc.get('scenario') == 'normal' for tc in test_cases),
                'has_param_test': any(tc.get('scenario', '').startswith('param') for tc in test_cases),
                'has_security': any(tc.get('scenario') == 'security' for tc in test_cases),
                'has_performance': any(tc.get('scenario') == 'performance' for tc in test_cases),
                'has_unauthorized': any(tc.get('scenario') == 'unauthorized' for tc in test_cases)
            }
        }

        return jsonify({'success': True, 'data': result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def parse_document(filepath: str, file_ext: str) -> Dict:
    """根据文件扩展名解析文档"""
    if file_ext in ['docx']:
        parser = WordParser()
        parsed = parser.parse(filepath)
        requirements = parser.extract_requirements(parsed.get('paragraphs', []))
        api_info = parser.extract_api_info(parsed.get('paragraphs', []))
    elif file_ext == 'pdf':
        parser = PDFParser()
        parsed = parser.parse(filepath)
        requirements = parser.extract_requirements(parsed.get('text', []))
        api_info = parser.extract_api_info(parsed.get('text', []))
    elif file_ext in ['md', 'txt']:
        parser = MarkdownParser()
        parsed = parser.parse(filepath)
        requirements = parser.extract_requirements(parsed.get('paragraphs', []), parsed.get('tables', []))
        api_info = parser.extract_api_info(parsed.get('paragraphs', []), parsed.get('code_blocks', []))
    else:
        return {'requirements': [], 'apiInfo': []}

    return {
        'requirements': requirements,
        'apiInfo': api_info,
        'headings': parsed.get('headings', []),
        'tables': parsed.get('tables', [])
    }


@web_bp.route('/api/encode', methods=['POST'])
def encode_text():
    data = request.get_json()
    text = data.get('text', '')
    encode_type = data.get('type', 'base64')

    try:
        result = ''
        if encode_type == 'base64':
            result = base64.b64encode(text.encode('utf-8')).decode('utf-8')
        elif encode_type == 'base64_decode':
            result = base64.b64decode(text).decode('utf-8')
        elif encode_type == 'url':
            result = urllib.parse.quote(text)
        elif encode_type == 'url_decode':
            result = urllib.parse.unquote(text)
        elif encode_type == 'md5':
            result = hashlib.md5(text.encode('utf-8')).hexdigest()
        elif encode_type == 'sha256':
            result = hashlib.sha256(text.encode('utf-8')).hexdigest()
        
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@web_bp.route('/api/text-process', methods=['POST'])
def text_process():
    data = request.get_json()
    text = data.get('text', '')
    process_type = data.get('type', 'upper')

    try:
        result = ''
        if process_type == 'upper':
            result = text.upper()
        elif process_type == 'lower':
            result = text.lower()
        elif process_type == 'trim':
            result = text.strip()
        elif process_type == 'reverse':
            result = text[::-1]
        elif process_type == 'word_count':
            chars = len(text)
            words = len(text.split())
            lines = text.count('\n') + 1
            result = f"字符数: {chars}\n词数: {words}\n行数: {lines}"
        elif process_type == 'remove_html':
            import re
            result = re.sub(r'<[^>]*>', '', text)
        
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@web_bp.route('/api/date-calc', methods=['POST'])
def date_calc():
    data = request.get_json()
    date_str = data.get('date', '')
    days = data.get('days', 0)
    calc_type = data.get('type', 'today')

    try:
        result = ''
        if calc_type == 'today':
            result = datetime.now().strftime('%Y-%m-%d')
        elif calc_type == 'timestamp':
            result = str(int(datetime.now().timestamp()))
        elif calc_type == 'add':
            if date_str:
                base_date = datetime.strptime(date_str, '%Y-%m-%d')
            else:
                base_date = datetime.now()
            result = (base_date + timedelta(days=days)).strftime('%Y-%m-%d')
        elif calc_type == 'subtract':
            if date_str:
                base_date = datetime.strptime(date_str, '%Y-%m-%d')
            else:
                base_date = datetime.now()
            result = (base_date - timedelta(days=days)).strftime('%Y-%m-%d')
        elif calc_type == 'weekday':
            if date_str:
                base_date = datetime.strptime(date_str, '%Y-%m-%d')
            else:
                base_date = datetime.now()
            weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
            result = weekdays[base_date.weekday()]
        elif calc_type == 'diff':
            if date_str:
                target_date = datetime.strptime(date_str, '%Y-%m-%d')
                today = datetime.now()
                diff = (target_date - today).days
                result = f"距离 {date_str} 还有 {abs(diff)} 天"
                if diff < 0:
                    result = f"{date_str} 已过去 {abs(diff)} 天"
        
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@web_bp.route('/api/network', methods=['POST'])
def network_tool():
    data = request.get_json()
    input_str = data.get('input', '')
    tool_type = data.get('type', 'url_parse')

    try:
        result = {}
        if tool_type == 'url_parse':
            parsed = urllib.parse.urlparse(input_str)
            result = {
                'scheme': parsed.scheme,
                'netloc': parsed.netloc,
                'path': parsed.path,
                'params': parsed.params,
                'query': parsed.query,
                'fragment': parsed.fragment
            }
        elif tool_type == 'url_encode':
            result = {'encoded': urllib.parse.quote(input_str)}
        elif tool_type == 'url_decode':
            result = {'decoded': urllib.parse.unquote(input_str)}
        elif tool_type == 'generate_url':
            protocols = ['http', 'https']
            domains = ['example.com', 'test.com', 'api.service.cn', 'dev.local']
            paths = ['/api/users', '/v1/login', '/data/query', '/service']
            import random
            result = {
                'url': f"{random.choice(protocols)}://{random.choice(domains)}{random.choice(paths)}?id={random.randint(1, 1000)}"
            }
        
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@web_bp.route('/api/log-analyze', methods=['POST'])
def log_analyze():
    """智能日志分析"""
    data = request.get_json()
    log_text = data.get('log', '')
    
    if not log_text:
        return jsonify({'success': False, 'error': '请输入日志内容'}), 400
    
    issues = []
    lines = log_text.split('\n')
    
    # 常见错误模式定义
    error_patterns = {
        '500': {
            'pattern': r'500|Internal Server Error|服务器内部错误',
            'level': 'error',
            'category': '服务器错误',
            'description': '服务器内部错误，通常是代码异常或配置问题',
            'causes': ['后端代码运行时异常未捕获', '数据库连接超时或连接池耗尽', '内存溢出或线程池耗尽', '第三方服务调用失败', '文件读写权限不足'],
            'solutions': ['查看错误堆栈定位具体代码位置', '检查数据库连接配置和连接池大小', '查看服务器内存和CPU使用率', '检查最近的代码发布和配置变更', '联系后端开发人员提供完整错误日志']
        },
        '502': {
            'pattern': r'502|Bad Gateway|网关错误',
            'level': 'error',
            'category': '网关错误',
            'description': '网关或代理服务器无法从上游服务器获取有效响应',
            'causes': ['上游服务未启动或崩溃', '上游服务响应超时', 'Nginx配置错误', '上游服务资源耗尽', '网络连接问题'],
            'solutions': ['检查上游服务是否正常运行', '查看上游服务日志', '检查Nginx配置', '调整超时时间', '检查网络连通性']
        },
        '504': {
            'pattern': r'504|Gateway Timeout|网关超时',
            'level': 'error',
            'category': '超时错误',
            'description': '网关等待上游服务器响应超时',
            'causes': ['上游服务处理时间过长', '数据库查询超时', '网关超时设置过短', '网络延迟', '服务负载过高'],
            'solutions': ['优化服务处理逻辑', '增加超时时间', '优化数据库查询', '使用异步处理', '增加服务实例']
        },
        '499': {
            'pattern': r'499|Client Closed Request',
            'level': 'warning',
            'category': '客户端断开',
            'description': '客户端在服务器响应前主动关闭了连接',
            'causes': ['用户等待时间过长', '网络不稳定断开', '请求超时', '页面被关闭'],
            'solutions': ['优化接口响应时间', '检查前端超时配置', '使用异步加载', '增加进度提示']
        },
        '400': {
            'pattern': r'400|Bad Request|参数错误',
            'level': 'error',
            'category': '请求参数错误',
            'description': '请求参数格式错误或缺少必填参数',
            'causes': ['参数格式不符合要求', '缺少必填参数', '参数值超出范围', 'Content-Type错误', 'JSON解析失败'],
            'solutions': ['检查参数格式', '确认必填参数', '检查参数值范围', '确认Content-Type', '验证JSON格式']
        },
        '401': {
            'pattern': r'401|Unauthorized|未授权|未登录',
            'level': 'warning',
            'category': '认证错误',
            'description': '请求缺少有效的认证信息',
            'causes': ['Token未携带或已过期', 'Token格式错误', '用户权限不足', '登录状态失效'],
            'solutions': ['检查Token', '确认Token有效期', '刷新Token或重新登录', '检查权限']
        },
        '403': {
            'pattern': r'403|Forbidden|禁止访问',
            'level': 'warning',
            'category': '权限错误',
            'description': '用户没有访问该资源的权限',
            'causes': ['角色权限不足', 'IP被限制', '访问频率超限', '访问白名单限制'],
            'solutions': ['联系管理员开通权限', '检查IP', '降低请求频率', '申请特殊权限']
        },
        '404': {
            'pattern': r'404|Not Found|找不到',
            'level': 'info',
            'category': '资源不存在',
            'description': '请求的资源不存在',
            'causes': ['URL路径错误', '资源已删除', '接口版本错误', '大小写不匹配'],
            'solutions': ['核对接口URL', '确认资源存在', '检查版本号', '检查大小写']
        },
        '503': {
            'pattern': r'503|Service Unavailable|服务不可用',
            'level': 'error',
            'category': '服务不可用',
            'description': '服务器暂时无法处理请求',
            'causes': ['服务器过载', '服务维护中', '服务被禁用', '第三方服务不可用'],
            'solutions': ['等待恢复', '查看服务状态', '检查计划维护', '确认依赖服务']
        },
        'NPE': {
            'pattern': r'NullPointerException|空指针|null is not',
            'level': 'error',
            'category': '空指针异常',
            'description': '代码尝试访问空对象的属性或方法',
            'causes': ['对象未初始化', '查询结果为null', '获取元素为null', '接口返回字段为null'],
            'solutions': ['进行null判断', '使用Optional', '确保字段有默认值', '查看堆栈定位']
        },
        'OOM': {
            'pattern': r'OutOfMemoryError|内存溢出|heap space',
            'level': 'error',
            'category': '内存溢出',
            'description': 'JVM堆内存不足',
            'causes': ['数据量过大', '内存泄漏', '并发过高', '堆内存配置过小'],
            'solutions': ['增加堆内存', '优化代码', '使用流式处理', '使用内存分析工具']
        },
        'SQL': {
            'pattern': r'SQLException|SQL error|数据库异常',
            'level': 'error',
            'category': '数据库错误',
            'description': 'SQL语句执行错误',
            'causes': ['SQL语法错误', '表字段不存在', '违反约束', '连接超时', '事务超时'],
            'solutions': ['检查SQL语法', '确认表结构', '检查数据约束', '检查连接池', '检查事务']
        },
        'Timeout': {
            'pattern': r'Timeout|超时|time out',
            'level': 'warning',
            'category': '超时错误',
            'description': '操作执行时间超出限制',
            'causes': ['数据库查询慢', '远程服务慢', '网络延迟', '负载过高'],
            'solutions': ['优化查询', '使用缓存', '增加超时', '增加资源']
        },
        'JSON': {
            'pattern': r'JSONException|JSON parse|解析失败',
            'level': 'error',
            'category': 'JSON解析错误',
            'description': 'JSON格式解析失败',
            'causes': ['JSON格式不规范', '包含非法字符', '编码问题', 'JSON被截断'],
            'solutions': ['验证JSON', '确保UTF-8编码', '检查特殊字符', '确认数据完整']
        },
        'Connection': {
            'pattern': r'Connection refused|连接失败|连接被拒绝',
            'level': 'error',
            'category': '连接错误',
            'description': '无法建立网络连接',
            'causes': ['服务未启动', '端口错误', '防火墙阻止', '网络不通'],
            'solutions': ['确认服务启动', '检查端口', '检查防火墙', '测试连通性']
        }
    }
    
    found_patterns = set()
    matched_issues = []
    
    for line_num, line in enumerate(lines, 1):
        for error_key, error_info in error_patterns.items():
            import re
            if re.search(error_info['pattern'], line, re.IGNORECASE):
                if error_key not in found_patterns:
                    found_patterns.add(error_key)
                    matched_issues.append({
                        'type': error_key,
                        'level': error_info['level'],
                        'category': error_info['category'],
                        'description': error_info['description'],
                        'causes': error_info['causes'],
                        'solutions': error_info['solutions'],
                        'matched_line': line.strip()[:200],
                        'line_number': line_num
                    })
    
    summary = {
        'total_errors': len([i for i in matched_issues if i['level'] == 'error']),
        'total_warnings': len([i for i in matched_issues if i['level'] == 'warning']),
        'total_info': len([i for i in matched_issues if i['level'] == 'info']),
        'unique_issues': len(matched_issues),
        'total_lines': len(lines)
    }
    
    return jsonify({'success': True, 'data': {'summary': summary, 'issues': matched_issues}})


@web_bp.route('/api/export-log-analysis', methods=['POST'])
def export_log_analysis():
    """导出日志分析结果为Excel"""
    data = request.get_json()
    issues = data.get('issues', [])
    
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '日志分析结果'
        
        # 表头样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        headers = ['序号', '错误等级', '问题类型', '描述', '匹配行', '行号', '可能原因', '排查建议']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # 错误等级颜色映射
        level_fill = {
            'error': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'warning': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
            'info': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        }
        
        # 填充数据
        for row_idx, issue in enumerate(issues, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            
            level_cell = ws.cell(row=row_idx, column=2, value=issue.get('level', ''))
            level_cell.fill = level_fill.get(issue.get('level'), PatternFill())
            
            ws.cell(row=row_idx, column=3, value=issue.get('category', ''))
            ws.cell(row=row_idx, column=4, value=issue.get('description', ''))
            ws.cell(row=row_idx, column=5, value=issue.get('matched_line', ''))
            ws.cell(row=row_idx, column=6, value=issue.get('line_number', ''))
            ws.cell(row=row_idx, column=7, value='\n'.join(issue.get('causes', [])))
            ws.cell(row=row_idx, column=8, value='\n'.join(issue.get('solutions', [])))
        
        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 30
        
        # 自动换行
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='日志分析结果.xlsx'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@web_bp.route('/api/format-json', methods=['POST'])
def format_json():
    """JSON/XML格式化"""
    data = request.get_json()
    content = data.get('content', '')
    format_type = data.get('type', 'json')
    
    if not content:
        return jsonify({'success': False, 'error': '请输入内容'}), 400
    
    try:
        if format_type == 'json':
            import json
            parsed = json.loads(content)
            formatted = json.dumps(parsed, ensure_ascii=False, indent=2)
            minified = json.dumps(parsed, separators=(',', ':'))
            return jsonify({'success': True, 'data': {'formatted': formatted, 'minified': minified, 'valid': True}})
        elif format_type == 'json_validate':
            import json
            json.loads(content)
            return jsonify({'success': True, 'data': {'valid': True, 'message': 'JSON格式正确'}})
        elif format_type == 'xml':
            import re
            formatted = re.sub(r'><', '>\n<', content)
            return jsonify({'success': True, 'data': {'formatted': formatted.strip(), 'valid': True}})
        elif format_type == 'sql':
            keywords = ['SELECT', 'FROM', 'WHERE', 'AND', 'OR', 'ORDER BY', 'GROUP BY', 'JOIN', 'LIMIT']
            formatted = content.upper()
            for kw in keywords:
                formatted = formatted.replace(kw, f'\n{kw}')
            return jsonify({'success': True, 'data': {'formatted': formatted.strip(), 'valid': True}})
        return jsonify({'success': True, 'data': {'formatted': content, 'valid': True}})
    except Exception as e:
        return jsonify({'success': False, 'error': f'格式化失败: {str(e)}'}), 400


@web_bp.route('/api/regex-test', methods=['POST'])
def regex_test():
    """正则表达式测试"""
    data = request.get_json()
    pattern = data.get('pattern', '')
    text = data.get('text', '')
    flags = data.get('flags', {'i': False, 'g': True})
    
    if not pattern:
        return jsonify({'success': False, 'error': '请输入正则表达式'}), 400
    
    try:
        import re
        flag = re.IGNORECASE if flags.get('i') else 0
        regex = re.compile(pattern, flag)
        matches = []
        for match in regex.finditer(text) if flags.get('g', True) else [regex.search(text)]:
            if match:
                matches.append({'match': match.group(), 'start': match.start(), 'end': match.end()})
        return jsonify({'success': True, 'data': {'matches': matches, 'count': len(matches), 'valid': True}})
    except re.error as e:
        return jsonify({'success': False, 'error': f'正则表达式错误: {str(e)}'}), 400


@web_bp.route('/api/timestamp-convert', methods=['POST'])
def timestamp_convert():
    """时间戳转换"""
    data = request.get_json()
    value = data.get('value', '')
    convert_type = data.get('type', 'to_date')
    
    try:
        from datetime import datetime
        if convert_type == 'to_date':
            timestamp = int(value)
            if timestamp > 10000000000:
                timestamp = timestamp // 1000
            dt = datetime.fromtimestamp(timestamp)
            weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
            return jsonify({'success': True, 'data': {
                'timestamp': timestamp, 'datetime': dt.strftime('%Y-%m-%d %H:%M:%S'),
                'date': dt.strftime('%Y-%m-%d'), 'time': dt.strftime('%H:%M:%S'), 'weekday': weekdays[dt.weekday()]
            }})
        else:
            dt = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
            return jsonify({'success': True, 'data': {
                'timestamp': int(dt.timestamp()), 'milliseconds': int(dt.timestamp() * 1000), 'datetime': value
            }})
    except Exception as e:
        return jsonify({'success': False, 'error': f'转换失败: {str(e)}'}), 400


@web_bp.route('/api/jwt-decode', methods=['POST'])
def jwt_decode():
    """JWT解码"""
    data = request.get_json()
    token = data.get('token', '').strip()
    
    if not token:
        return jsonify({'success': False, 'error': '请输入JWT Token'}), 400
    
    try:
        parts = token.split('.')
        if len(parts) != 3:
            return jsonify({'success': False, 'error': 'JWT格式错误'}), 400
        
        import base64, json
        from datetime import datetime
        
        def decode_base64url(data):
            padding = 4 - len(data) % 4
            if padding != 4:
                data += '=' * padding
            return json.loads(base64.urlsafe_b64decode(data))
        
        header = decode_base64url(parts[0])
        payload = decode_base64url(parts[1])
        
        if 'exp' in payload:
            payload['exp_readable'] = datetime.fromtimestamp(payload['exp']).strftime('%Y-%m-%d %H:%M:%S')
        if 'iat' in payload:
            payload['iat_readable'] = datetime.fromtimestamp(payload['iat']).strftime('%Y-%m-%d %H:%M:%S')
        
        return jsonify({'success': True, 'data': {
            'header': header, 'payload': payload, 'signature': parts[2][:20] + '...', 'valid': True
        }})
    except Exception as e:
        return jsonify({'success': False, 'error': f'解码失败: {str(e)}'}), 400


@web_bp.route('/api/hash-generate', methods=['POST'])
def hash_generate():
    """生成哈希值"""
    data = request.get_json()
    text = data.get('text', '')
    hash_type = data.get('type', 'md5')
    
    if not text:
        return jsonify({'success': False, 'error': '请输入文本'}), 400
    
    try:
        text_bytes = text.encode('utf-8')
        result = {}
        if hash_type in ['md5', 'all']:
            result['md5'] = hashlib.md5(text_bytes).hexdigest()
        if hash_type in ['sha1', 'all']:
            result['sha1'] = hashlib.sha1(text_bytes).hexdigest()
        if hash_type in ['sha256', 'all']:
            result['sha256'] = hashlib.sha256(text_bytes).hexdigest()
        if hash_type in ['sha512', 'all']:
            result['sha512'] = hashlib.sha512(text_bytes).hexdigest()
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@web_bp.route('/api/color-convert', methods=['POST'])
def color_convert():
    """颜色转换"""
    data = request.get_json()
    color = data.get('color', '')
    convert_type = data.get('type', 'hex_to_rgb')
    
    try:
        import re
        if convert_type == 'hex_to_rgb':
            hex_color = color.lstrip('#')
            if len(hex_color) == 3:
                hex_color = ''.join([c*2 for c in hex_color])
            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
            return jsonify({'success': True, 'data': {'hex': f'#{hex_color.upper()}', 'rgb': f'rgb({r}, {g}, {b})', 'r': r, 'g': g, 'b': b}})
        else:
            match = re.match(r'rgb\((\d+),\s*(\d+),\s*(\d+)\)', color)
            if not match:
                return jsonify({'success': False, 'error': 'RGB格式错误'}), 400
            r, g, b = int(match.group(1)), int(match.group(2)), int(match.group(3))
            return jsonify({'success': True, 'data': {'hex': f'#{r:02X}{g:02X}{b:02X}', 'rgb': color, 'r': r, 'g': g, 'b': b}})
    except Exception as e:
        return jsonify({'success': False, 'error': f'转换失败: {str(e)}'}), 400


@web_bp.route('/api/generate-uuid', methods=['POST'])
def generate_uuid():
    """生成UUID"""
    import uuid
    data = request.get_json()
    count = min(data.get('count', 1), 100)
    uuids = [str(uuid.uuid4()) for _ in range(count)]
    return jsonify({'success': True, 'data': uuids})


@web_bp.route('/api/parse-xmind', methods=['POST'])
def parse_xmind():
    """解析XMind文件 - 兼容XMind 8 / 2020+ / 2023+ 多种格式"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '请上传XMind文件'}), 400
    
    file = request.files['file']
    if not file.filename.endswith('.xmind'):
        return jsonify({'success': False, 'error': '请上传XMind格式文件'}), 400
    
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        import tempfile
        import os
        import re
        
        temp_dir = tempfile.gettempdir()
        xmind_path = os.path.join(temp_dir, file.filename)
        file.save(xmind_path)
        
        test_cases = []
        NS = 'urn:xmind:xmap:xmlns:content:2.0'
        
        with zipfile.ZipFile(xmind_path, 'r') as zf:
            # 尝试多个可能的content文件路径
            content_paths = ['content.xml', 'content.json', 'Content.json']
            xml_content = None
            for cp in content_paths:
                if cp in zf.namelist():
                    xml_content = zf.read(cp)
                    break
            
            if xml_content is None:
                # 尝试找任何.xml文件
                xml_files = [n for n in zf.namelist() if n.endswith('.xml')]
                for xf in xml_files:
                    xml_content = zf.read(xf)
                    break
            
            if xml_content is None:
                os.remove(xmind_path)
                return jsonify({'success': False, 'error': '无法找到XMind内容文件，请确认文件有效'}), 400
            
            root = ET.fromstring(xml_content)
            
            # 通用函数：按标签局部名查找元素（忽略命名空间）
            def find_tags(el, local_name):
                if el.tag.endswith('}' + local_name):
                    return True
                return el.tag == local_name
            
            # 获取所有sheet或topic根节点
            def get_all_sheets(root_elem):
                sheets = []
                # 尝试标准命名空间
                for sheet in root_elem.findall(f'.//{{{NS}}}sheet'):
                    sheets.append(sheet)
                # 尝试无命名空间
                if not sheets:
                    for el in root_elem.iter():
                        if el.tag.endswith('}sheet') or el.tag == 'sheet':
                            sheets.append(el)
                # 尝试直接把根元素当sheet
                if not sheets:
                    for el in root_elem.iter():
                        if el.tag.endswith('}topic') and el.find(f'./{{{NS}}}title') is not None:
                            # 直接把根topic当sheet处理
                            fake_sheet = ET.Element('sheet')
                            fake_sheet.append(el)
                            sheets.append(fake_sheet)
                            break
                return sheets
            
            def get_text(el, tag_local_name):
                """从元素中获取指定标签的文本（兼容命名空间）"""
                ns_tag = f'{{{NS}}}{tag_local_name}'
                child = el.find(ns_tag)
                if child is not None:
                    return child.text or ''
                # 尝试无命名空间
                for c in el:
                    if c.tag.endswith('}' + tag_local_name) or c.tag == tag_local_name:
                        return c.text or ''
                return ''
            
            def get_children(el, tag_local_name='topic'):
                """获取指定标签名的子元素（兼容命名空间）"""
                ns_tag = f'{{{NS}}}{tag_local_name}'
                children = el.findall(ns_tag)
                if children:
                    return children
                # 尝试通过topics包装层
                topics_wrapper = el.find(f'{{{NS}}}topics')
                if topics_wrapper is not None:
                    children = topics_wrapper.findall(ns_tag)
                    if children:
                        return children
                # 尝试无命名空间
                result = []
                for c in el:
                    if c.tag.endswith('}' + tag_local_name) or c.tag == tag_local_name:
                        result.append(c)
                    elif c.tag.endswith('}topics') or c.tag == 'topics':
                        for cc in c:
                            if cc.tag.endswith('}' + tag_local_name) or cc.tag == tag_local_name:
                                result.append(cc)
                return result
            
            def extract_case_from_topic(topic, depth=0, path=''):
                """从topic节点递归提取测试用例"""
                title = get_text(topic, 'title')
                if not title:
                    return
                
                current_path = path + ' / ' + title if path else title
                children = get_children(topic)
                
                # 判断当前节点是否为测试用例节点
                title_lower = title.lower()
                
                # 检查是否有步骤/预期等子节点标识（深度>=1的节点）
                has_step_children = False
                for child in children:
                    child_title = get_text(child, 'title').lower()
                    if any(kw in child_title for kw in ['步骤', 'step', '预期', 'expected', '期望', '优先级', 'priority', '前置', 'precondition']):
                        has_step_children = True
                        break
                
                # 是测试用例的条件
                is_test_case = False
                if depth >= 1 and children and has_step_children:
                    is_test_case = True
                elif depth >= 1 and title and any(kw in title_lower for kw in ['用例', '测试', 'tc-', 'test', 'case']):
                    is_test_case = True
                # 深度>=2且有子节点也当作用例
                elif depth >= 2 and len(children) >= 1:
                    is_test_case = True
                
                if is_test_case:
                    # 创建新用例
                    case_id = ''
                    case_name = title
                    
                    id_match = re.search(r'(TC-\d+|用例\d+|ID:\s*\d+)', title)
                    if id_match:
                        case_id = id_match.group(1).replace('ID:', '').strip()
                        case_name = title.replace(id_match.group(1), '').strip()
                    
                    new_case = {
                        'id': case_id,
                        'name': case_name or title,
                        'description': '',
                        'priority': 'P2',
                        'type': '功能测试',
                        'test_steps': [],
                        'test_data': {},
                        'expected': '',
                        'preconditions': ''
                    }
                    
                    test_cases.append(new_case)
                    
                    # 递归处理子节点作为用例属性
                    for child in children:
                        parse_case_attributes(child, new_case)
                    return new_case
                else:
                    # 非用例节点，继续向下查找
                    for child in children:
                        extract_case_from_topic(child, depth + 1, current_path)
            
            def parse_case_attributes(topic, current_case):
                """解析测试用例的属性（步骤、预期、优先级等）"""
                title = get_text(topic, 'title')
                if not title:
                    return
                
                title_lower = title.lower()
                children = get_children(topic)
                
                # 优先级
                if any(kw in title_lower for kw in ['优先级', 'priority', 'p0', 'p1', 'p2', '高', '中', '低']):
                    priority_map = {
                        'p0': 'P0', '高': 'P0', '最高': 'P0',
                        'p1': 'P1', '中': 'P1', '中等': 'P1',
                        'p2': 'P2', '低': 'P2', '最低': 'P2'
                    }
                    val = title_lower.strip()
                    for k, v in priority_map.items():
                        if k in val:
                            current_case['priority'] = v
                            break
                    return
                
                # 类型
                if any(kw in title_lower for kw in ['类型', 'type']):
                    clean = title.replace('类型', '').replace('type', '').strip()
                    type_map = {'功能': '功能测试', '接口': '接口测试', '性能': '性能测试', '安全': '安全测试'}
                    for k, v in type_map.items():
                        if k in clean.lower():
                            current_case['type'] = v
                            return
                    return
                
                # 描述
                if any(kw in title_lower for kw in ['描述', 'description', '说明', '概述']):
                    for kw in ['描述', 'description', '说明', '概述']:
                        title = title.replace(kw, '').strip()
                    current_case['description'] = title
                    return
                
                # 前置条件
                if any(kw in title_lower for kw in ['前置条件', 'precondition', '前提条件', '前提']):
                    for kw in ['前置条件', 'precondition', '前提条件', '前提']:
                        title = title.replace(kw, '').strip()
                    current_case['preconditions'] = title
                    return
                
                # 预期结果
                if any(kw in title_lower for kw in ['预期', 'expected', '期望结果', '预期结果']):
                    for kw in ['预期结果', '期望结果', 'expected', '预期', '期望']:
                        title = title.replace(kw, '').strip()
                    title = title.lstrip('：:').strip()
                    current_case['expected'] = title
                    return
                
                # 测试数据
                if any(kw in title_lower for kw in ['测试数据', '测试参数', '数据', 'data', '参数']):
                    current_case['test_data'] = {'title': title}
                    return
                
                # 步骤解析（包含数字或步骤关键字）
                is_step = False
                if any(kw in title_lower for kw in ['步骤', 'step', '操作', 'action']):
                    is_step = True
                if re.search(r'^\d+[.、.。]?\s*', title):
                    is_step = True
                # 剩余的子节点也作为步骤
                if not is_step and not children:
                    is_step = True
                
                if is_step:
                    step_match = re.match(r'(\d+)[.、.。]?\s*(.*)', title)
                    if step_match:
                        step_num = int(step_match.group(1))
                        step_desc = step_match.group(2).strip()
                    else:
                        step_num = len(current_case['test_steps']) + 1
                        step_desc = title
                    # 清理步骤文本中的"步骤N："前缀
                    step_desc = re.sub(r'^(步骤|step|操作|action)\s*\d*\s*[：:,.，。]?\s*', '', step_desc, flags=re.IGNORECASE).strip()
                    if not step_desc:
                        step_desc = title
                    
                    # 递归检查子节点是否为预期结果
                    step_expected = ''
                    for child in children:
                        ct = get_text(child, 'title').lower()
                        if any(kw in ct for kw in ['预期', 'expected', '期望']):
                            step_expected = get_text(child, 'title')
                            for kw in ['预期', '期望', 'expected']:
                                step_expected = step_expected.replace(kw, '').strip() if kw in step_expected else step_expected
                            break
                    
                    current_case['test_steps'].append({
                        'step': step_num,
                        'action': step_desc,
                        'expected': step_expected
                    })
                    
                    # 继续解析其他子节点
                    for child in children:
                        parse_case_attributes(child, current_case)
                    return
                
                # 默认：继续递归解析子节点
                for child in children:
                    parse_case_attributes(child, current_case)
            
            # 主流程：从所有sheet中提取用例
            sheets = get_all_sheets(root)
            for sheet in sheets:
                # 先尝试找topic根节点
                topics = get_children(sheet)
                for topic in topics:
                    extract_case_from_topic(topic, 0)
        
        os.remove(xmind_path)
        
        # 为没有ID的用例生成ID
        for i, tc in enumerate(test_cases, 1):
            if not tc.get('id'):
                tc['id'] = f'TC-{i:04d}'
        
        return jsonify({'success': True, 'data': test_cases})
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@web_bp.route('/api/parse-apk', methods=['POST'])
def parse_apk():
    """解析APK文件"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '请上传APK文件'}), 400
    
    file = request.files['file']
    if not file.filename.endswith('.apk'):
        return jsonify({'success': False, 'error': '请上传APK格式文件'}), 400
    
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        import hashlib
        import tempfile
        import os
        
        # 使用系统兼容的临时目录
        temp_dir = tempfile.gettempdir()
        apk_path = os.path.join(temp_dir, file.filename)
        file.save(apk_path)
        
        apk_info = {
            'filename': file.filename,
            'package_name': '',
            'version_name': '',
            'version_code': '',
            'min_sdk': '',
            'target_sdk': '',
            'permissions': [],
            'features': [],
            'md5': '',
            'size': 0
        }
        
        # 获取文件大小
        import os
        apk_info['size'] = os.path.getsize(apk_path)
        
        # 计算MD5
        with open(apk_path, 'rb') as f:
            apk_info['md5'] = hashlib.md5(f.read()).hexdigest()
        
        # 解析APK（ZIP格式）
        with zipfile.ZipFile(apk_path, 'r') as zf:
            # 解析AndroidManifest.xml
            if 'AndroidManifest.xml' in zf.namelist():
                manifest_data = zf.read('AndroidManifest.xml')
                
                # AndroidManifest.xml是二进制格式，尝试解析
                try:
                    # 尝试直接解析XML
                    tree = ET.fromstring(manifest_data)
                    ns = {'android': 'http://schemas.android.com/apk/res/android'}
                    
                    apk_info['package_name'] = tree.get('package', '')
                    apk_info['version_name'] = tree.get('{%s}versionName' % ns['android'], '')
                    apk_info['version_code'] = tree.get('{%s}versionCode' % ns['android'], '')
                    
                    # 获取SDK版本
                    uses_sdk = tree.find('.//uses-sdk')
                    if uses_sdk is not None:
                        apk_info['min_sdk'] = uses_sdk.get('{%s}minSdkVersion' % ns['android'], '')
                        apk_info['target_sdk'] = uses_sdk.get('{%s}targetSdkVersion' % ns['android'], '')
                    
                    # 获取权限
                    for permission in tree.findall('.//uses-permission'):
                        perm = permission.get('{%s}name' % ns['android'], '')
                        if perm:
                            apk_info['permissions'].append(perm)
                    
                    # 获取features
                    for feature in tree.findall('.//uses-feature'):
                        feat = feature.get('{%s}name' % ns['android'], '')
                        if feat:
                            apk_info['features'].append(feat)
                except Exception:
                    # 二进制XML解析失败，使用aapt工具
                    import subprocess
                    try:
                        result = subprocess.run(
                            ['aapt', 'dump', 'badging', apk_path],
                            capture_output=True, text=True, timeout=60
                        )
                        if result.returncode == 0:
                            output = result.stdout
                            # 解析aapt输出
                            for line in output.split('\n'):
                                if line.startswith('package:'):
                                    import re
                                    pkg_match = re.search(r"name='([^']+)'", line)
                                    ver_name_match = re.search(r"versionName='([^']+)'", line)
                                    ver_code_match = re.search(r"versionCode='([^']+)'", line)
                                    if pkg_match:
                                        apk_info['package_name'] = pkg_match.group(1)
                                    if ver_name_match:
                                        apk_info['version_name'] = ver_name_match.group(1)
                                    if ver_code_match:
                                        apk_info['version_code'] = ver_code_match.group(1)
                                elif line.startswith('uses-permission:'):
                                    perm_match = re.search(r"name='([^']+)'", line)
                                    if perm_match:
                                        apk_info['permissions'].append(perm_match.group(1))
                                elif line.startswith('sdkVersion:'):
                                    apk_info['min_sdk'] = line.split("'")[1]
                                elif line.startswith('targetSdkVersion:'):
                                    apk_info['target_sdk'] = line.split("'")[1]
                    except:
                        pass
        
        # 清理临时文件
        os.remove(apk_path)
        
        # 格式化大小
        size = apk_info['size']
        if size < 1024:
            apk_info['size_display'] = f"{size} B"
        elif size < 1024 * 1024:
            apk_info['size_display'] = f"{size / 1024:.2f} KB"
        else:
            apk_info['size_display'] = f"{size / (1024 * 1024):.2f} MB"
        
        return jsonify({'success': True, 'data': apk_info})
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'解析失败: {str(e)}'}), 500


# ==================== 常用命令模块 ====================

SQL_COMMANDS = {
    '基础查询': [
        {'command': 'SELECT', 'syntax': 'SELECT 列 FROM 表', 'example': 'SELECT id, name FROM users;', 'desc': '从表中查询数据'},
        {'command': 'WHERE', 'syntax': 'SELECT ... WHERE 条件', 'example': "SELECT * FROM users WHERE age > 18;", 'desc': '按条件筛选数据'},
        {'command': 'ORDER BY', 'syntax': 'SELECT ... ORDER BY 列 [ASC|DESC]', 'example': 'SELECT * FROM users ORDER BY created_at DESC;', 'desc': '对查询结果排序'},
        {'command': 'GROUP BY', 'syntax': 'SELECT 列, 聚合函数 FROM 表 GROUP BY 列', 'example': 'SELECT city, COUNT(*) FROM users GROUP BY city;', 'desc': '按列分组聚合'},
        {'command': 'LIMIT', 'syntax': 'SELECT ... LIMIT 数量 OFFSET 偏移', 'example': 'SELECT * FROM users LIMIT 10 OFFSET 20;', 'desc': '限制返回行数（分页）'},
        {'command': 'DISTINCT', 'syntax': 'SELECT DISTINCT 列 FROM 表', 'example': 'SELECT DISTINCT city FROM users;', 'desc': '去重查询'},
        {'command': 'IN', 'syntax': 'SELECT ... WHERE 列 IN (值)', 'example': 'SELECT * FROM users WHERE id IN (1,2,3);', 'desc': '匹配值列表'},
        {'command': 'BETWEEN', 'syntax': 'SELECT ... WHERE 列 BETWEEN a AND b', 'example': 'SELECT * FROM users WHERE age BETWEEN 18 AND 30;', 'desc': '范围查询'},
        {'command': 'LIKE', 'syntax': 'SELECT ... WHERE 列 LIKE 模式', 'example': "SELECT * FROM users WHERE name LIKE '张%';", 'desc': '模糊匹配'},
    ],
    '连接与子查询': [
        {'command': 'INNER JOIN', 'syntax': '表1 JOIN 表2 ON 条件', 'example': 'SELECT u.name, o.amount FROM users u JOIN orders o ON u.id = o.user_id;', 'desc': '内连接'},
        {'command': 'LEFT JOIN', 'syntax': '表1 LEFT JOIN 表2 ON 条件', 'example': 'SELECT u.name, o.amount FROM users u LEFT JOIN orders o ON u.id = o.user_id;', 'desc': '左连接（保留左表全部）'},
        {'command': 'RIGHT JOIN', 'syntax': '表1 RIGHT JOIN 表2 ON 条件', 'example': 'SELECT * FROM users u RIGHT JOIN orders o ON u.id = o.user_id;', 'desc': '右连接'},
        {'command': 'UNION', 'syntax': '查询1 UNION 查询2', 'example': 'SELECT id FROM users UNION SELECT id FROM orders;', 'desc': '合并查询结果并去重'},
        {'command': '子查询', 'syntax': 'WHERE 列 IN (SELECT ...)', 'example': "SELECT * FROM users WHERE id IN (SELECT user_id FROM orders);", 'desc': '嵌套查询'},
    ],
    '数据定义与操纵': [
        {'command': 'CREATE TABLE', 'syntax': 'CREATE TABLE 表 (列 类型, ...)', 'example': 'CREATE TABLE users (id INT PRIMARY KEY, name VARCHAR(50));', 'desc': '创建表'},
        {'command': 'ALTER TABLE', 'syntax': 'ALTER TABLE 表 ADD/MODIFY 列', 'example': 'ALTER TABLE users ADD COLUMN email VARCHAR(100);', 'desc': '修改表结构'},
        {'command': 'DROP TABLE', 'syntax': 'DROP TABLE 表名', 'example': 'DROP TABLE test_users;', 'desc': '删除表（谨慎）'},
        {'command': 'INSERT', 'syntax': 'INSERT INTO 表 (列) VALUES (值)', 'example': "INSERT INTO users (name, age) VALUES ('张三', 25);", 'desc': '插入记录'},
        {'command': 'UPDATE', 'syntax': 'UPDATE 表 SET 列=值 WHERE 条件', 'example': "UPDATE users SET age = 26 WHERE id = 1;", 'desc': '更新记录（务必加WHERE）'},
        {'command': 'DELETE', 'syntax': 'DELETE FROM 表 WHERE 条件', 'example': 'DELETE FROM users WHERE id = 999;', 'desc': '删除记录（务必加WHERE）'},
        {'command': 'TRUNCATE', 'syntax': 'TRUNCATE TABLE 表名', 'example': 'TRUNCATE TABLE test_users;', 'desc': '清空表（不可回滚）'},
        {'command': 'CREATE INDEX', 'syntax': 'CREATE INDEX 索引名 ON 表(列)', 'example': 'CREATE INDEX idx_email ON users(email);', 'desc': '创建索引加速查询'},
    ],
    '函数与高级': [
        {'command': 'COUNT', 'syntax': 'SELECT COUNT(*) FROM 表', 'example': 'SELECT COUNT(*) FROM users;', 'desc': '统计行数'},
        {'command': 'SUM / AVG', 'syntax': 'SUM(列) / AVG(列)', 'example': 'SELECT SUM(amount), AVG(amount) FROM orders;', 'desc': '求和与平均值'},
        {'command': 'MAX / MIN', 'syntax': 'MAX(列) / MIN(列)', 'example': 'SELECT MAX(age), MIN(age) FROM users;', 'desc': '求最大值与最小值'},
        {'command': 'HAVING', 'syntax': 'GROUP BY ... HAVING 条件', 'example': 'SELECT city, COUNT(*) FROM users GROUP BY city HAVING COUNT(*) > 10;', 'desc': '对聚合结果筛选'},
        {'command': 'EXPLAIN', 'syntax': 'EXPLAIN SELECT ...', 'example': "EXPLAIN SELECT * FROM users WHERE email = 'x@x.com';", 'desc': '查看查询执行计划'},
        {'command': 'CASE WHEN', 'syntax': 'CASE WHEN 条件 THEN 值 ELSE 值 END', 'example': "SELECT name, CASE WHEN age>=18 THEN '成年' ELSE '未成年' END AS status FROM users;", 'desc': '条件表达式'},
        {'command': '事务', 'syntax': 'BEGIN; ... COMMIT / ROLLBACK;', 'example': 'BEGIN; UPDATE accounts SET money=money-100 WHERE id=1; COMMIT;', 'desc': '事务（原子操作）'},
    ],
}

LINUX_COMMANDS = {
    '文件目录操作': [
        {'command': 'ls', 'syntax': 'ls [-l -a -h] [路径]', 'example': 'ls -lah /home/user', 'desc': '列出目录内容'},
        {'command': 'cd', 'syntax': 'cd 路径', 'example': 'cd /var/log', 'desc': '切换目录'},
        {'command': 'pwd', 'syntax': 'pwd', 'example': 'pwd', 'desc': '显示当前路径'},
        {'command': 'cp', 'syntax': 'cp [-r] 源 目标', 'example': 'cp -r src/ backup/', 'desc': '复制文件/目录'},
        {'command': 'mv', 'syntax': 'mv 源 目标', 'example': 'mv a.txt b.txt', 'desc': '移动或重命名'},
        {'command': 'rm', 'syntax': 'rm [-r -f] 文件', 'example': 'rm -rf temp/', 'desc': '删除文件（-r递归 -f强制）'},
        {'command': 'mkdir', 'syntax': 'mkdir [-p] 目录', 'example': 'mkdir -p a/b/c', 'desc': '创建目录（-p递归创建）'},
        {'command': 'find', 'syntax': 'find 路径 -name "模式"', 'example': 'find /tmp -name "*.log" -mtime -7', 'desc': '按名称/时间查找文件'},
    ],
    '文本内容查看': [
        {'command': 'cat', 'syntax': 'cat 文件', 'example': 'cat /etc/hosts', 'desc': '查看小文件内容'},
        {'command': 'head', 'syntax': 'head -n N 文件', 'example': 'head -20 app.log', 'desc': '查看文件前N行'},
        {'command': 'tail', 'syntax': 'tail [-n N -f] 文件', 'example': 'tail -f app.log | grep ERROR', 'desc': '查看末尾 / 实时追踪日志'},
        {'command': 'less', 'syntax': 'less 文件', 'example': 'less /var/log/syslog', 'desc': '分页查看大文件（q退出 /搜索）'},
        {'command': 'grep', 'syntax': 'grep [选项] 模式 文件', 'example': 'grep -n "ERROR" app.log', 'desc': '搜索文本（-n行号 -i忽略大小写）'},
        {'command': 'wc', 'syntax': 'wc [-l -w -c] 文件', 'example': 'wc -l app.log', 'desc': '统计行数/单词数'},
        {'command': 'sort / uniq', 'syntax': 'sort | uniq -c', 'example': 'cat log | sort | uniq -c | sort -rn', 'desc': '排序去重并统计出现次数'},
        {'command': 'awk', 'syntax': "awk '{print $N}' 文件", 'example': "awk '{print $1}' access.log | sort | uniq -c", 'desc': '按列提取文本'},
    ],
    '系统进程监控': [
        {'command': 'top', 'syntax': 'top / htop', 'example': 'top', 'desc': '实时进程与资源占用（q退出）'},
        {'command': 'ps', 'syntax': 'ps [-aux -ef]', 'example': 'ps aux | grep java', 'desc': '查看运行进程'},
        {'command': 'kill', 'syntax': 'kill [-9] PID', 'example': 'kill -9 12345', 'desc': '结束进程（-9强制杀死）'},
        {'command': 'df', 'syntax': 'df -h', 'example': 'df -h', 'desc': '查看磁盘空间'},
        {'command': 'du', 'syntax': 'du -sh [路径]', 'example': 'du -sh /var/log/*', 'desc': '查看目录/文件大小'},
        {'command': 'free', 'syntax': 'free [-h -m -g]', 'example': 'free -g', 'desc': '查看内存使用'},
        {'command': 'lsof', 'syntax': 'lsof -i:端口', 'example': 'lsof -i:8080', 'desc': '查看端口占用进程'},
    ],
    '网络与权限': [
        {'command': 'curl', 'syntax': 'curl [选项] URL', 'example': 'curl -i http://api.test.com/users', 'desc': 'HTTP请求（接口调试）'},
        {'command': 'wget', 'syntax': 'wget [-O 新名] URL', 'example': 'wget http://test.com/file.zip', 'desc': '下载文件'},
        {'command': 'ping', 'syntax': 'ping 主机 [-c N]', 'example': 'ping baidu.com -c 4', 'desc': '测试网络连通性'},
        {'command': 'ssh', 'syntax': 'ssh 用户@主机', 'example': 'ssh root@192.168.1.100', 'desc': '远程登录服务器'},
        {'command': 'chmod', 'syntax': 'chmod 权限 文件', 'example': 'chmod 755 script.sh', 'desc': '修改文件权限（r=4 w=2 x=1）'},
        {'command': 'chown', 'syntax': 'chown 用户:组 文件', 'example': 'chown -R app:app /data/', 'desc': '修改文件所有者'},
        {'command': 'systemctl', 'syntax': 'systemctl start/stop/status 服务', 'example': 'systemctl restart nginx', 'desc': '管理系统服务'},
    ],
}

ADB_COMMANDS = {
    '设备连接管理': [
        {'command': 'adb devices', 'syntax': 'adb devices', 'example': 'adb devices -l', 'desc': '列出已连接设备'},
        {'command': 'adb connect', 'syntax': 'adb connect IP:端口', 'example': 'adb connect 192.168.1.100:5555', 'desc': 'WiFi连接设备'},
        {'command': 'adb disconnect', 'syntax': 'adb disconnect [IP]', 'example': 'adb disconnect 192.168.1.100', 'desc': '断开WiFi连接'},
        {'command': 'adb -s', 'syntax': 'adb -s 设备ID 命令', 'example': 'adb -s emulator-5554 shell', 'desc': '多设备时指定设备'},
        {'command': 'adb kill-server', 'syntax': 'adb kill-server && adb start-server', 'example': 'adb kill-server && adb start-server', 'desc': '重启ADB服务'},
        {'command': 'adb version', 'syntax': 'adb version', 'example': 'adb version', 'desc': '查看ADB版本'},
    ],
    '应用安装管理': [
        {'command': 'adb install', 'syntax': 'adb install [-r -t -d] APK', 'example': 'adb install -r app-debug.apk', 'desc': '安装APK（-r覆盖 -t测试包）'},
        {'command': 'adb uninstall', 'syntax': 'adb uninstall 包名', 'example': 'adb uninstall com.example.app', 'desc': '卸载应用'},
        {'command': 'pm list packages', 'syntax': 'adb shell pm list packages [-3 -s]', 'example': 'adb shell pm list packages -3', 'desc': '列出已安装包（-3第三方）'},
        {'command': 'pm clear', 'syntax': 'adb shell pm clear 包名', 'example': 'adb shell pm clear com.example.app', 'desc': '清除应用数据'},
        {'command': 'am start', 'syntax': 'adb shell am start -n 包名/Activity', 'example': 'adb shell am start -n com.example.app/.MainActivity', 'desc': '启动指定Activity'},
        {'command': 'am force-stop', 'syntax': 'adb shell am force-stop 包名', 'example': 'adb shell am force-stop com.example.app', 'desc': '强制停止应用'},
    ],
    '日志与截图': [
        {'command': 'adb logcat', 'syntax': 'adb logcat [选项]', 'example': 'adb logcat -v time -d > crash.log', 'desc': '查看Android日志'},
        {'command': 'logcat 过滤', 'syntax': 'adb logcat TAG:LEVEL *:E', 'example': 'adb logcat MyApp:D *:E', 'desc': '按标签/级别过滤日志'},
        {'command': 'logcat grep', 'syntax': 'adb logcat | grep 关键词', 'example': 'adb logcat | grep -i "crash"', 'desc': '只看关键词日志'},
        {'command': 'logcat -c', 'syntax': 'adb logcat -c && adb logcat', 'example': 'adb logcat -c && adb logcat -v time', 'desc': '清空缓冲区后重录日志'},
        {'command': 'screencap', 'syntax': 'adb exec-out screencap -p > 文件', 'example': 'adb exec-out screencap -p > bug.png', 'desc': '截取屏幕保存到本地'},
        {'command': 'screenrecord', 'syntax': 'adb shell screenrecord [时长] 路径', 'example': 'adb shell screenrecord --time-limit 30 /sdcard/bug.mp4', 'desc': '录屏（复现Bug）'},
    ],
    '模拟输入操作': [
        {'command': 'input tap', 'syntax': 'adb shell input tap X Y', 'example': 'adb shell input tap 500 800', 'desc': '点击屏幕坐标'},
        {'command': 'input swipe', 'syntax': 'adb shell input swipe X1 Y1 X2 Y2 [ms]', 'example': 'adb shell input swipe 100 800 800 800 300', 'desc': '滑动屏幕'},
        {'command': 'input text', 'syntax': 'adb shell input text "字符串"', 'example': 'adb shell input text "hello world"', 'desc': '输入文本（空格用%s）'},
        {'command': 'input keyevent', 'syntax': 'adb shell input keyevent 键码', 'example': 'adb shell input keyevent 4（返回键）', 'desc': '模拟按键（3=HOME 4=BACK 26=电源）'},
    ],
    '系统调试与性能': [
        {'command': 'adb shell', 'syntax': 'adb shell', 'example': 'adb shell', 'desc': '进入设备Shell命令行'},
        {'command': 'adb dumpsys', 'syntax': 'adb shell dumpsys [子系统]', 'example': 'adb shell dumpsys meminfo com.example.app', 'desc': '查看内存/Activity等系统状态'},
        {'command': 'adb shell top', 'syntax': 'adb shell top [-m N]', 'example': 'adb shell top -m 10 -d 2 -s cpu', 'desc': '查看进程CPU/内存占用'},
        {'command': 'adb reboot', 'syntax': 'adb reboot [bootloader]', 'example': 'adb reboot bootloader', 'desc': '重启 / 进入Fastboot'},
        {'command': 'getprop', 'syntax': 'adb shell getprop [键]', 'example': 'adb shell getprop ro.build.version.release', 'desc': '获取系统属性（Android版本等）'},
        {'command': 'bugreport', 'syntax': 'adb bugreport [路径]', 'example': 'adb bugreport ./bugreport.zip', 'desc': '导出完整Bug报告'},
    ],
}


@web_bp.route('/api/commands', methods=['POST'])
def get_commands():
    """获取常用命令列表（sql / linux / adb）"""
    try:
        data = request.get_json() or {}
        cmd_type = data.get('type', 'sql')
        
        mapping = {
            'sql': SQL_COMMANDS,
            'linux': LINUX_COMMANDS,
            'adb': ADB_COMMANDS
        }
        
        commands = mapping.get(cmd_type)
        if commands is None:
            return jsonify({'success': False, 'error': f'不支持的命令类型: {cmd_type}'}), 400
        
        # 按分类组织返回
        categories = []
        for cat_name, cmd_list in commands.items():
            categories.append({
                'category': cat_name,
                'commands': cmd_list
            })
        
        return jsonify({'success': True, 'data': categories})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@web_bp.route('/api/commands/search', methods=['POST'])
def search_commands():
    """搜索命令（在所有三类命令中搜索关键字）"""
    try:
        data = request.get_json() or {}
        keyword = (data.get('keyword') or '').strip().lower()
        cmd_type = data.get('type', '')  # 可选：限定某一类
        
        if not keyword:
            return jsonify({'success': True, 'data': []})
        
        mapping = {
            'sql': SQL_COMMANDS,
            'linux': LINUX_COMMANDS,
            'adb': ADB_COMMANDS
        }
        
        sources = [(cmd_type, mapping[cmd_type])] if cmd_type in mapping else list(mapping.items())
        
        results = []
        for t_name, cmd_map in sources:
            for cat_name, cmd_list in cmd_map.items():
                for cmd in cmd_list:
                    text = f"{cmd['command']} {cmd.get('desc', '')} {cmd.get('example', '')}".lower()
                    if keyword in text:
                        cmd_copy = dict(cmd)
                        cmd_copy['type'] = t_name
                        cmd_copy['category'] = cat_name
                        results.append(cmd_copy)
        
        return jsonify({'success': True, 'data': results[:100]})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
